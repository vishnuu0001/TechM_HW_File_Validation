import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_validation_report(input_path, output_path):
    """
    Optimized version: Reads the input Excel file, validates 'Compute' tab against 'README-Glossary',
    and saves the report to output_path using vectorized operations.
    """
    try:
        # 1. Load README-Glossary with only required columns
        df_glossary = pd.read_excel(
            input_path, 
            sheet_name='README-Glossary', 
            header=6,
            usecols=['Tab Name', 'Column Name']
        )

        # Clean Glossary Data
        df_glossary['Tab Name'] = df_glossary['Tab Name'].astype(str).str.strip()
        df_glossary['Column Name'] = df_glossary['Column Name'].astype(str).str.strip()

        # Get valid columns for 'Compute' tab as a set for O(1) lookup
        valid_compute_columns = set(
            df_glossary.loc[df_glossary['Tab Name'] == 'Compute', 'Column Name'].values
        )

        # 2. Load only necessary columns from Compute Sheet
        df_compute = pd.read_excel(input_path, sheet_name='Compute', header=5)

        # Clean column names once
        df_compute.columns = df_compute.columns.astype(str).str.strip()

        # 3. Get column indices and names
        columns = df_compute.columns.tolist()
        
        # Column indices (ensure they exist)
        if len(columns) <= 23:
            return False, "Compute sheet doesn't have enough columns.", None
        
        # Define column positions
        idx_sbg = 2
        idx_ban = 3
        idx_app_name = 4
        idx_server_id = 13
        idx_sep_scenario = 17
        target_indices = [18, 19, 20, 21, 22, 23]

        # Get column names for target indices
        target_columns = [columns[i] for i in target_indices if i < len(columns)]
        
        # Filter target columns that exist in glossary
        valid_target_columns = [col for col in target_columns if col in valid_compute_columns]
        
        if not valid_target_columns:
            return False, "No valid target columns found in glossary.", None

        # 4. Vectorized filtering: Get rows where separation scenario is "TBD"
        sep_col_name = columns[idx_sep_scenario]
        mask_tbd = df_compute[sep_col_name].astype(str).str.strip().str.upper() == 'TBD'
        df_tbd = df_compute[mask_tbd].copy()

        if df_tbd.empty:
            return False, "No records found with 'TBD' in Server-Level Separation Scenario.", None

        # 5. Vectorized missing column detection
        def get_missing_columns(row):
            missing = []
            for col in valid_target_columns:
                val = row[col]
                if pd.isna(val) or str(val).strip() == "":
                    missing.append(col)
            return "\n".join(missing) if missing else None

        # Apply missing column check
        df_tbd['Columns Missing'] = df_tbd.apply(get_missing_columns, axis=1)
        
        # Filter only rows with missing columns
        df_filtered = df_tbd[df_tbd['Columns Missing'].notna()].copy()

        if df_filtered.empty:
            return False, "No records found with missing data in target columns.", None

        # 6. Create report DataFrame using vectorized operations
        report_df = pd.DataFrame({
            "Business Application Number (BAN)": df_filtered.iloc[:, idx_ban].fillna("N/A"),
            "Category": "Compute",
            "SBG": df_filtered.iloc[:, idx_sbg].fillna("N/A"),
            "Business Application Name": df_filtered.iloc[:, idx_app_name].fillna("N/A"),
            "Server ID / Name": df_filtered.iloc[:, idx_server_id].fillna("N/A"),
            "Server-Level Separation Scenario": df_filtered.iloc[:, idx_sep_scenario].fillna("N/A"),
            "Columns Missing": df_filtered['Columns Missing']
        })

        # 7. Write to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Compute')

        # 8. Apply formatting
        apply_formatting(output_path, 'Compute', len(report_df))

        # 9. Calculate statistics
        stats = calculate_statistics(report_df)

        return True, f"Generated {len(report_df)} records.", stats

    except Exception as e:
        return False, str(e), None


def apply_formatting(file_path, sheet_name, row_count):
    """
    Optimized: Apply conditional formatting, alignment, and styling to the Excel report.
    """
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # Define styles (reuse objects for performance) - All using Aptos font size 10
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Aptos", bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        tbd_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        tbd_font = Font(name="Aptos", bold=True, color="C65911", size=10)
        
        missing_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        missing_font = Font(name="Aptos", color="C00000", size=10)
        
        # Default font for regular cells
        default_font = Font(name="Aptos", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Set column widths (batch operation)
        column_widths = [25, 12, 15, 35, 25, 30, 50]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Format header row
        ws.row_dimensions[1].height = 40
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Pre-calculate which columns need which alignment
        center_cols = {1, 2, 3, 6}  # A, B, C, F
        left_cols = {4, 5}  # D, E
        missing_col = 7  # G
        
        # Format data rows - optimized loop
        for row_idx in range(2, row_count + 2):
            row_fill = light_fill if row_idx % 2 == 0 else white_fill
            row_height = 30  # Default height
            
            for col_idx in range(1, 8):  # 7 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                if col_idx in center_cols:
                    cell.alignment = center_alignment
                    # Special handling for TBD in column F
                    if col_idx == 6 and str(cell.value).strip().upper() == 'TBD':
                        cell.fill = tbd_fill
                        cell.font = tbd_font
                    else:
                        cell.fill = row_fill
                        cell.font = default_font
                elif col_idx in left_cols:
                    cell.alignment = left_alignment
                    cell.fill = row_fill
                    cell.font = default_font
                elif col_idx == missing_col:
                    cell.alignment = left_alignment
                    cell.fill = missing_fill
                    cell.font = missing_font
                    # Calculate row height based on line breaks
                    if cell.value:
                        line_count = str(cell.value).count('\n') + 1
                        if line_count > 1:
                            row_height = max(15 * line_count, 30)
            
            # Set row height once per row
            ws.row_dimensions[row_idx].height = row_height
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save the workbook
        wb.save(file_path)
        
    except Exception as e:
        print(f"Error applying formatting: {str(e)}")
        # Don't fail the whole process if formatting fails


def calculate_statistics(df_report):
    """
    Calculate statistics from the validation report.
    Returns a dictionary with summary statistics including category breakdown.
    """
    try:
        stats = {
            'total_records': len(df_report),
            'unique_sbg_count': df_report['SBG'].nunique(),
            'unique_ban_count': df_report['Business Application Number (BAN)'].nunique(),
            'unique_categories': df_report['Category'].nunique(),
            'sbg_list': df_report['SBG'].unique().tolist(),
            'category_list': df_report['Category'].unique().tolist(),
            'sbg_breakdown': df_report.groupby('SBG').size().to_dict(),
            'ban_breakdown': df_report.groupby('Business Application Number (BAN)').size().to_dict(),
            'category_breakdown': df_report.groupby('Category').size().to_dict(),
            'category_details': {}
        }
        
        # For each category, get distinct SBGs and their BAN counts
        for category in df_report['Category'].unique():
            df_category = df_report[df_report['Category'] == category]
            distinct_sbgs = df_category['SBG'].unique().tolist()
            
            sbg_ban_details = {}
            for sbg in distinct_sbgs:
                df_sbg = df_category[df_category['SBG'] == sbg]
                distinct_bans = df_sbg['Business Application Number (BAN)'].nunique()
                sbg_ban_details[sbg] = {
                    'distinct_bans': distinct_bans,
                    'total_records': len(df_sbg)
                }
            
            stats['category_details'][category] = {
                'distinct_sbgs': len(distinct_sbgs),
                'sbg_ban_details': sbg_ban_details
            }
        
        return stats
    except Exception as e:
        print(f"Error calculating statistics: {str(e)}")
        return {
            'total_records': 0,
            'unique_sbg_count': 0,
            'unique_ban_count': 0,
            'unique_categories': 0,
            'sbg_list': [],
            'category_list': [],
            'sbg_breakdown': {},
            'ban_breakdown': {},
            'category_breakdown': {},
            'category_details': {}
        }